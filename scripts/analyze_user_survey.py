#!/usr/bin/env python3
from __future__ import annotations

import csv
import html
import json
import re
import textwrap
from collections import Counter
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "doc/survey/yandex-responses-2026-05-12.xlsx"
OUT = ROOT / "doc/survey"


def read_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v) if v is not None else "" for v in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        data.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return headers, data


def num(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    m = re.match(r"^([1-5])", text)
    return int(m.group(1)) if m else None


def count_column(rows: list[dict[str, object]], column: str) -> Counter[str]:
    c: Counter[str] = Counter()
    for row in rows:
        value = row.get(column)
        if value is not None and str(value).strip():
            c[str(value).strip()] += 1
    return c


def multi_counts(rows: list[dict[str, object]], prefix: str) -> Counter[str]:
    c: Counter[str] = Counter()
    columns = [key for key in rows[0].keys() if key.startswith(prefix + " / ")]
    for row in rows:
        for col in columns:
            value = row.get(col)
            if value is not None and str(value).strip():
                label = col.split(" / ", 1)[1]
                c[label] += 1
    return c


def rating_summary(rows: list[dict[str, object]], columns: dict[str, str]) -> list[dict[str, object]]:
    result = []
    for label, col in columns.items():
        values = [num(row.get(col)) for row in rows]
        values = [v for v in values if v is not None]
        if not values:
            continue
        result.append(
            {
                "label": label,
                "avg": round(mean(values), 2),
                "top2": round(sum(1 for v in values if v >= 4) * 100 / len(values)),
                "count": len(values),
            }
        )
    return sorted(result, key=lambda item: (item["top2"], item["avg"]), reverse=True)


def pct(value: int, total: int) -> str:
    return f"{round(value * 100 / total)}%" if total else "0%"


def table(headers: list[str], rows: list[list[object]]) -> str:
    out = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        out.append("| " + " | ".join(str(x) for x in row) + " |")
    return "\n".join(out)


def wrapped_text(label: str, x: int, y: int, cls: str, max_chars: int = 36) -> str:
    lines = textwrap.wrap(label, width=max_chars) or [label]
    tspans = []
    for i, line in enumerate(lines[:3]):
        dy = 0 if i == 0 else 24
        tspans.append(f'<tspan x="{x}" dy="{dy}">{html.escape(line)}</tspan>')
    return f'<text x="{x}" y="{y}" class="{cls}">{"".join(tspans)}</text>'


def bar_svg(title: str, data: list[tuple[str, int, str]], path: Path, width: int = 1800) -> None:
    row_h = 78
    top = 96
    left = 620
    right = 170
    height = top + row_h * len(data) + 48
    max_value = max((value for _, value, _ in data), default=1)
    bar_w = width - left - right
    rows = []
    for i, (label, value, suffix) in enumerate(data):
        y = top + i * row_h
        w = int(bar_w * value / max_value) if max_value else 0
        rows.append(
            f"""
  {wrapped_text(label, 48, y + 24, "label")}
  <rect x="{left}" y="{y}" width="{w}" height="34" rx="11" fill="#3f7cff"/>
  <text x="{left + w + 12}" y="{y + 25}" class="value">{html.escape(str(value) + suffix)}</text>"""
        )
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <defs>
    <style>
      .title {{ font-family: Arial, sans-serif; font-size: 34px; font-weight: 700; fill: #18202e; }}
      .label {{ font-family: Arial, sans-serif; font-size: 21px; fill: #253047; }}
      .value {{ font-family: Arial, sans-serif; font-size: 22px; font-weight: 700; fill: #18202e; }}
    </style>
  </defs>
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="48" y="54" class="title">{html.escape(title)}</text>
{''.join(rows)}
</svg>
"""
    path.write_text(svg, encoding="utf-8")


def grouped_bar_svg(title: str, groups: list[tuple[str, Counter[str]]], path: Path) -> None:
    labels = ["0–20% времени", "21–40% времени", "41–60% времени", "61–80% времени", "81–100% времени"]
    width = 1200
    height = 430
    left = 210
    top = 110
    group_h = 110
    bar_w = 760
    colors = ["#ef6a6a", "#f2a65a", "#f2d45a", "#74bf78", "#3f9f68"]
    rows = []
    for gi, (name, counts) in enumerate(groups):
        y = top + gi * group_h
        total = sum(counts.values()) or 1
        x = left
        rows.append(f'<text x="48" y="{y + 31}" class="label">{html.escape(name)}</text>')
        for li, label in enumerate(labels):
            value = counts.get(label, 0)
            w = int(bar_w * value / total)
            if w:
                rows.append(f'<rect x="{x}" y="{y}" width="{w}" height="42" fill="{colors[li]}"/>')
                if w > 44:
                    rows.append(f'<text x="{x + w / 2:.0f}" y="{y + 28}" text-anchor="middle" class="inside">{value}</text>')
            x += w
    legend = []
    x = left
    for color, label in zip(colors, labels):
        legend.append(f'<rect x="{x}" y="340" width="18" height="18" fill="{color}"/><text x="{x + 25}" y="356" class="legend">{html.escape(label)}</text>')
        x += 180
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <defs>
    <style>
      .title {{ font-family: Arial, sans-serif; font-size: 34px; font-weight: 700; fill: #18202e; }}
      .label {{ font-family: Arial, sans-serif; font-size: 23px; font-weight: 700; fill: #253047; }}
      .inside {{ font-family: Arial, sans-serif; font-size: 20px; font-weight: 700; fill: #ffffff; }}
      .legend {{ font-family: Arial, sans-serif; font-size: 18px; fill: #253047; }}
    </style>
  </defs>
  <rect width="100%" height="100%" fill="#ffffff"/>
  <text x="48" y="54" class="title">{html.escape(title)}</text>
{''.join(rows)}
{''.join(legend)}
</svg>
"""
    path.write_text(svg, encoding="utf-8")


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def main() -> None:
    headers, rows = read_rows(INPUT)
    if not rows:
        raise SystemExit("No responses")
    total = len(rows)

    role_col = "1. Какая роль вам ближе?"
    create_col = "4. Создаёте ли вы маршруты самостоятельно?"
    internet_col = "14. Какую долю времени в типичном маршруте у вас доступен стабильный мобильный интернет?"
    gps_col = "15. Какую долю времени GPS на смартфоне определяется корректно и без заметных проблем?"

    roles = count_column(rows, role_col)
    creators = count_column(rows, create_col)
    problems = multi_counts(rows, "8. С какими проблемами вы сталкиваетесь чаще всего при работе с маршрутами?")
    tools = multi_counts(rows, "5. Какими инструментами вы сейчас пользуетесь для работы с маршрутами?")
    route_types = multi_counts(rows, "2. Какой тип маршрутов вам наиболее интересен?")
    internet = count_column(rows, internet_col)
    gps = count_column(rows, gps_col)

    rating_columns = {
        "Фото в контексте маршрута": "9. Насколько для вас важна возможность видеть фотографии прямо в контексте маршрута на карте? (1 — не важно, 5 — очень важно)",
        "Публикация по ссылке": "10. Насколько для вас важна возможность публиковать маршрут по ссылке для других пользователей? (1 — не важно, 5 — очень важно)",
        "Социальные функции": "11. Насколько для вас важны социальные функции маршрутов: комментарии, лайки, рейтинги, закладки? (1 — не важно, 5 — очень важно)",
        "Маршрут при слабом интернете": "16. Насколько для вас важна возможность открыть маршрут при слабом интернете или без него? (1 — не важно, 5 — очень важно)",
        "Положение относительно маршрута": "17. Насколько для вас важна возможность видеть своё текущее положение относительно маршрута? (1 — не важно, 5 — очень важно)",
        "Автозапись маршрута": "19. Насколько для вас важна возможность автоматически записывать фактический маршрут во время движения? (1 — не важно, 5 — очень важно)",
        "Текст к точке": "20. Насколько для вас важна возможность добавлять к точке не только фото, но и текстовый комментарий или описание? (1 — не важно, 5 — очень важно)",
        "Дата и время прохождения": "22. Насколько для вас полезна возможность указать дату и время прохождения маршрута? (1 — не полезно, 5 — очень полезно)",
        "Фото и комментарии во время прохождения": "24. Насколько вам полезны фото и комментарии к отдельным точкам маршрута во время прохождения? (1 — не полезно, 5 — очень полезно)",
    }
    ratings = rating_summary(rows, rating_columns)

    function_columns = {
        "Создание на карте": "12. Оцените важность следующих функций будущей системы / создание маршрута на интерактивной карте",
        "Фото к точкам": "12. Оцените важность следующих функций будущей системы / добавление фотографий к точкам",
        "Заметки к точкам": "12. Оцените важность следующих функций будущей системы / текстовые заметки к точкам",
        "Расстояние и время": "12. Оцените важность следующих функций будущей системы / расчёт расстояния и времени прохождения",
        "Профиль высот": "12. Оцените важность следующих функций будущей системы / просмотр профиля высот",
        "Поиск и фильтрация": "12. Оцените важность следующих функций будущей системы / поиск и фильтрация маршрутов",
        "Экспорт GPX/KML": "12. Оцените важность следующих функций будущей системы / экспорт в GPX/KML",
        "Публикация по ссылке": "12. Оцените важность следующих функций будущей системы / публикация маршрута по ссылке",
        "Просмотр без авторизации": "12. Оцените важность следующих функций будущей системы / просмотр маршрута без авторизации",
        "Офлайн-доступ": "12. Оцените важность следующих функций будущей системы / офлайн-доступ к ранее открытым маршрутам",
        "Погода": "12. Оцените важность следующих функций будущей системы / прогноз погоды для маршрута",
        "AI-ассистент": "12. Оцените важность следующих функций будущей системы / помощь ИИ-ассистента при поиске и построении маршрута",
    }
    functions = rating_summary(rows, function_columns)

    role_rows = [[k, v, pct(v, total)] for k, v in roles.most_common()]
    problem_rows = [[k, v, pct(v, total)] for k, v in problems.most_common()]
    rating_rows = [[r["label"], r["avg"], f'{r["top2"]}%', r["count"]] for r in ratings]
    function_rows = [[r["label"], r["avg"], f'{r["top2"]}%', r["count"]] for r in functions]

    write_csv(OUT / "summary-roles.csv", ["Группа", "Количество", "Доля"], role_rows)
    write_csv(OUT / "summary-problems.csv", ["Проблема", "Количество", "Доля"], problem_rows)
    write_csv(OUT / "summary-ratings.csv", ["Функция", "Средняя оценка", "Доля оценок 4-5", "Ответов"], rating_rows)
    write_csv(OUT / "summary-functions.csv", ["Функция", "Средняя оценка", "Доля оценок 4-5", "Ответов"], function_rows)

    charts = OUT / "charts"
    charts.mkdir(exist_ok=True)
    bar_svg(
        "Основные проблемы существующих инструментов",
        [(k, v, f" / {pct(v, total)}") for k, v in problems.most_common(8)],
        charts / "survey-problems.svg",
    )
    bar_svg(
        "Наиболее значимые функции системы",
        [(str(r["label"]), int(r["top2"]), "%") for r in functions[:8]],
        charts / "survey-functions.svg",
    )
    bar_svg(
        "Ключевые оценки пользователей",
        [(str(r["label"]), int(r["top2"]), "%") for r in ratings[:8]],
        charts / "survey-key-ratings.svg",
    )
    grouped_bar_svg("Доступность интернета и GPS в маршрутах", [("Интернет", internet), ("GPS", gps)], charts / "survey-connectivity.svg")

    authors = sum(1 for row in rows if str(row.get(create_col, "")).strip() in {"да, регулярно", "иногда"})
    viewers = total - authors
    low_internet = sum(internet.get(x, 0) for x in ["0–20% времени", "21–40% времени", "41–60% времени"])
    high_photo = next((r for r in ratings if r["label"] == "Фото в контексте маршрута"), None)
    high_link = next((r for r in ratings if r["label"] == "Публикация по ссылке"), None)
    high_offline = next((r for r in ratings if r["label"] == "Маршрут при слабом интернете"), None)

    md = f"""# Анализ результатов пилотного опроса

Дата выгрузки: 12.05.2026.

Файл исходных ответов: `yandex-responses-2026-05-12.xlsx`.

В опросе приняли участие {total} респондентов. Опрос рассматривается как пилотный и не претендует на статистическую репрезентативность, но используется для проверки того, совпадают ли выявленные в работе проблемы с опытом потенциальных пользователей.

## Профиль респондентов

{table(["Группа", "Количество", "Доля"], role_rows)}

Создают маршруты регулярно или иногда: {authors} из {total} ({pct(authors, total)}). В основном только проходят готовые маршруты: {viewers} из {total} ({pct(viewers, total)}).

## Интересующие типы маршрутов

{table(["Тип маршрута", "Количество", "Доля"], [[k, v, pct(v, total)] for k, v in route_types.most_common()])}

## Используемые инструменты

{table(["Инструмент", "Количество", "Доля"], [[k, v, pct(v, total)] for k, v in tools.most_common()])}

## Основные проблемы

{table(["Проблема", "Количество", "Доля"], problem_rows)}

## Значимость ключевых функций

{table(["Функция", "Средняя оценка", "Доля оценок 4-5", "Ответов"], rating_rows)}

## Оценка функций будущей системы

{table(["Функция", "Средняя оценка", "Доля оценок 4-5", "Ответов"], function_rows)}

## Интернет и GPS

{table(["Доступность интернета", "Количество", "Доля"], [[k, v, pct(v, total)] for k, v in internet.items()])}

{table(["Качество GPS", "Количество", "Доля"], [[k, v, pct(v, total)] for k, v in gps.items()])}

Стабильный мобильный интернет доступен не более 60% времени у {low_internet} из {total} респондентов ({pct(low_internet, total)}).

## Диаграммы

- `charts/survey-problems.svg` — основные проблемы существующих инструментов.
- `charts/survey-functions.svg` — наиболее значимые функции системы.
- `charts/survey-key-ratings.svg` — ключевые оценки пользователей.
- `charts/survey-connectivity.svg` — доступность интернета и GPS.

PNG-версии диаграмм лежат в той же папке, если выполнена конвертация через `rsvg-convert`.

## Готовый вывод для ВКР

Для уточнения потребностей целевой аудитории был проведён пилотный опрос потенциальных пользователей системы. В опросе приняли участие {total} респондентов, среди которых были как авторы маршрутов, так и пользователи, проходящие готовые маршруты. Результаты не претендуют на статистическую репрезентативность, однако позволяют сопоставить выявленные проблемы с пользовательским опытом.

Наиболее часто отмеченными проблемами стали: {', '.join(k for k, _ in problems.most_common(3))}. Возможность видеть фотографии в контексте маршрута получила долю высоких оценок {high_photo["top2"] if high_photo else 0}%, публикация маршрута по ссылке — {high_link["top2"] if high_link else 0}%, а открытие маршрута при слабом интернете — {high_offline["top2"] if high_offline else 0}%. Также {pct(low_internet, total)} респондентов указали, что стабильный мобильный интернет доступен не более 60% времени типичного маршрута.

Полученные результаты подтверждают актуальность функций, связанных с объединением маршрута, фотографий, описаний, публикации и мобильного использования в одном рабочем процессе. Поэтому результаты опроса были использованы как дополнительное обоснование требований к системе Guide Helper.

## Готовый короткий ответ комиссии

Опрос был пилотным, N={total}. Я не использую его как единственное доказательство актуальности, но он подтверждает выбранные функции: пользователям важны фотографии в контексте маршрута, публикация по ссылке, понятная карта и доступ к маршруту в условиях нестабильного интернета. Основная доказательная база работы — анализ аналогов, требования, реализация и тестирование, а опрос дополняет её пользовательским контекстом.
"""
    (OUT / "analysis.md").write_text(md, encoding="utf-8")

    summary = {
        "responses": total,
        "authors": authors,
        "viewers": viewers,
        "low_internet_count": low_internet,
        "roles": dict(roles),
        "problems": dict(problems),
        "ratings": ratings,
        "functions": functions,
        "internet": dict(internet),
        "gps": dict(gps),
    }
    (OUT / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
